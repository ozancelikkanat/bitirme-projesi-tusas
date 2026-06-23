"""Streamlit interface for the composite wrinkle risk model."""

from __future__ import annotations

from io import BytesIO, StringIO

import pandas as pd
import numpy as np
import streamlit as st
from openpyxl.styles import Font, PatternFill

from wrinkle_life_risk.config import ColumnMap, RiskThresholds
from wrinkle_life_risk.data_cleaning import MissingColumnError
from wrinkle_life_risk.ods_io import read_ods, read_ods_sheet_names
from wrinkle_life_risk.pipeline import AnalysisOptions, run_analysis
from wrinkle_life_risk.risk_labels import normalize_risk_label as normalize_risk_label_core
from wrinkle_life_risk.sample_data import make_sample_dataframe
from wrinkle_life_risk.visualization import (
    plot_delta_k_distribution,
    plot_remaining_life_distribution,
    plot_risk_class_counts,
    plot_risk_distribution,
    plot_theta_vs_measure,
    plot_top_critical_defects,
)


st.set_page_config(page_title="Kompozit Wrinkle Risk Analizi", layout="wide")


DISPLAY_COLUMN_NAMES = {
    "Wrinkle_ID": "Wrinkle ID",
    "defect_id": "Kusur ID",
    "Row_ID": "Satır ID",
    "Section": "Bölüm",
    "Category": "Kategori",
    "Length_mm": "Uzunluk (mm)",
    "Width_mm": "Genişlik (mm)",
    "Depth_mm": "Derinlik (mm)",
    "Height_mm": "Yükseklik (mm)",
    "effective_measure": "Etkin Ölçü",
    "effective_measure_source": "Etkin Ölçü Kaynağı",
    "theta_degree": "Açısal Sapma (derece)",
    "Severity_Index": "Severity Index",
    "Risk_Skor": "Severity Index",
    "risk_score": "Severity Index",
    "risk_class": "Geometric_Risk",
    "Risk_Level": "Final_Risk_Level",
    "Final Risk_Skor": "Final_Risk_Level",
    "Final_Risk_Level": "Final_Risk_Level",
    "Delta_sigma_MPa": "Delta sigma (MPa)",
    "Y_factor": "Y faktörü",
    "a_mm": "a (mm)",
    "a_m": "a (m)",
    "Delta_K_MPa_sqrt_m": "Delta K (MPa√m)",
    "Assumed_Stress_MPa": "Kabul edilen gerilme (MPa)",
    "Assumed_Crack_mm": "Eşdeğer çatlak boyutu (mm)",
    "Assumed_Crack_m": "Eşdeğer çatlak boyutu (m)",
    "Estimated_Life_Cycles": "Paris ters-hız göstergesi",
    "relative_life_index": "Göreceli ömür göstergesi (0–100)",
    "Paris_C": "Paris C",
    "Paris_m": "Paris m",
    "da_dN_m_per_cycle": "da/dN (m/cycle)",
    "critical_crack_size_mm": "Kritik çatlak boyutu (mm)",
    "remaining_life_cycles": "Kalan çevrim",
    "inspection_priority": "Inspection Priority",
    "fracture_status": "Kırılma durumu",
    "yorum_özeti": "Yorum özeti",
    "geometrik_değerlendirme": "Geometrik Değerlendirme",
    "kırılma_mekaniği_değerlendirmesi": "Kırılma Mekaniği Değerlendirmesi",
    "ömür_öncelik_yorumu": "Ömür / Öncelik Yorumu",
    "data_quality_status": "Veri Durumu",
}

RESULT_COLUMNS = [
    "Wrinkle_ID",
    "Row_ID",
    "Length_mm",
    "Width_mm",
    "Depth_mm",
    "Height_mm",
    "effective_measure",
    "effective_measure_source",
    "theta_degree",
    "Severity_Index",
    "Risk_Skor",
    "risk_score",
    "risk_class",
    "Delta_sigma_MPa",
    "Y_factor",
    "a_mm",
    "a_m",
    "Assumed_Stress_MPa",
    "Assumed_Crack_mm",
    "Assumed_Crack_m",
    "Delta_K_MPa_sqrt_m",
    "Paris_C",
    "Paris_m",
    "da_dN_m_per_cycle",
    "critical_crack_size_mm",
    "remaining_life_cycles",
    "Estimated_Life_Cycles",
    "relative_life_index",
    "inspection_priority",
    "fracture_status",
    "yorum_özeti",
]

PRIMARY_RESULT_COLUMNS = [
    "Wrinkle_ID",
    "Row_ID",
    "Section",
    "Category",
    "theta_degree",
    "Severity_Index",
    "Risk_Skor",
    "risk_score",
    "risk_class",
    "Final Risk_Skor",
    "Length_mm",
    "Width_mm",
    "Depth_mm",
    "Height_mm",
    "effective_measure",
    "effective_measure_source",
]

DETAIL_ONLY_COLUMNS = {
    "geometrik_değerlendirme",
    "kırılma_mekaniği_değerlendirmesi",
    "ömür_öncelik_yorumu",
    "engineering_comment",
}

RISK_DISPLAY_LABELS = {
    "CRITICAL": "Kritik",
    "HIGH": "Yüksek",
    "MEDIUM": "Orta",
    "LOW": "Düşük",
    "INCOMPLETE": "Hesap dışı",
    "High": "Yüksek",
    "Medium": "Orta",
    "Low": "Düşük",
    "Incomplete": "Hesap dışı",
}

RISK_CELL_COLORS = {
    "CRITICAL": "background-color: #fde2e2; color: #7f1d1d;",
    "HIGH": "background-color: #ffedd5; color: #7c2d12;",
    "MEDIUM": "background-color: #fff3bf; color: #713f12;",
    "LOW": "background-color: #dcfce7; color: #14532d;",
    "INCOMPLETE": "background-color: #e5e7eb; color: #374151;",
}

RISK_EXCEL_FILLS = {
    "CRITICAL": PatternFill("solid", fgColor="FDE2E2"),
    "HIGH": PatternFill("solid", fgColor="FFEDD5"),
    "MEDIUM": PatternFill("solid", fgColor="FFF3BF"),
    "LOW": PatternFill("solid", fgColor="DCFCE7"),
    "INCOMPLETE": PatternFill("solid", fgColor="E5E7EB"),
}


def inject_app_styles() -> None:
    st.markdown(
        """
        <style>
        .main .block-container {
            padding-top: 2rem;
            max-width: 1380px;
        }
        div[data-testid="stMetric"] {
            background: #ffffff;
            border: 1px solid #e5e7eb;
            border-radius: 8px;
            padding: 14px 16px;
            box-shadow: 0 1px 2px rgba(15, 23, 42, 0.05);
        }
        div[data-testid="stMetricLabel"] {
            color: #475569;
        }
        .hero-panel {
            border: 1px solid #dbe3ef;
            border-radius: 8px;
            padding: 22px 24px;
            background: linear-gradient(135deg, #f8fafc 0%, #eef6ff 100%);
            margin-bottom: 18px;
        }
        .hero-kicker {
            color: #005f8f;
            font-size: 0.82rem;
            font-weight: 700;
            letter-spacing: 0.06em;
            text-transform: uppercase;
            margin-bottom: 6px;
        }
        .hero-title {
            color: #111827;
            font-size: 2.15rem;
            font-weight: 760;
            line-height: 1.15;
            margin-bottom: 8px;
        }
        .hero-copy {
            color: #475569;
            font-size: 1rem;
            max-width: 980px;
        }
        .flow-grid {
            display: grid;
            grid-template-columns: repeat(4, minmax(0, 1fr));
            gap: 10px;
            margin: 14px 0 18px 0;
        }
        .flow-card {
            background: #ffffff;
            border: 1px solid #dbe3ef;
            border-radius: 8px;
            padding: 12px 14px;
            min-height: 98px;
        }
        .flow-step {
            color: #64748b;
            font-size: 0.76rem;
            font-weight: 700;
            text-transform: uppercase;
            margin-bottom: 6px;
        }
        .flow-title {
            color: #0f172a;
            font-weight: 720;
            margin-bottom: 4px;
        }
        .flow-text {
            color: #64748b;
            font-size: 0.86rem;
            line-height: 1.35;
        }
        .risk-strip {
            display: flex;
            gap: 10px;
            flex-wrap: wrap;
            margin: 8px 0 14px 0;
        }
        .risk-pill {
            border-radius: 999px;
            padding: 7px 12px;
            font-size: 0.86rem;
            font-weight: 700;
            border: 1px solid rgba(15, 23, 42, 0.08);
        }
        .risk-low { background: #dcfce7; color: #14532d; }
        .risk-medium { background: #fff3bf; color: #713f12; }
        .risk-high { background: #ffedd5; color: #7c2d12; }
        .risk-critical { background: #fde2e2; color: #7f1d1d; }
        .risk-note { background: #e0f2fe; color: #075985; }
        .section-note {
            border-left: 4px solid #0ea5e9;
            background: #f0f9ff;
            color: #0f172a;
            padding: 10px 12px;
            border-radius: 6px;
            margin: 8px 0 14px 0;
        }
        @media (max-width: 900px) {
            .flow-grid { grid-template-columns: 1fr; }
            .hero-title { font-size: 1.65rem; }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_hero() -> None:
    st.markdown(
        """
        <div class="hero-panel">
            <div class="hero-kicker">Composite Wrinkle Decision Support</div>
            <div class="hero-title">Kompozit Wrinkle Risk Analizi</div>
            <div class="hero-copy">
                Wrinkle kusurlarını iki katmanda değerlendirir: önce geometrik severity ve risk sınıfı,
                ardından yalnızca Orta, Yüksek ve Kritik kusurlar için ΔK, Paris Yasası ve göreceli ömür göstergesi.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_model_flow() -> None:
    st.markdown(
        """
        <div class="flow-grid">
            <div class="flow-card">
                <div class="flow-step">Katman 1</div>
                <div class="flow-title">Geometri taraması</div>
                <div class="flow-text">Genişlik, derinlik/yükseklik ve etkin ölçü ile açısal sapma hesaplanır.</div>
            </div>
            <div class="flow-card">
                <div class="flow-step">Risk eşiği</div>
                <div class="flow-title">Düşük / Orta / Yüksek / Kritik</div>
                <div class="flow-text">Risk skoru sabit eşiklerle sınıflandırılır; Düşük sınıf ileri analize aktarılmaz.</div>
            </div>
            <div class="flow-card">
                <div class="flow-step">Katman 2</div>
                <div class="flow-title">Kırılma mekaniği</div>
                <div class="flow-text">Orta, Yüksek ve Kritik kusurlar için ΔK, da/dN ve göreceli ömür göstergesi üretilir.</div>
            </div>
            <div class="flow-card">
                <div class="flow-step">Çıktı</div>
                <div class="flow-title">İnceleme önceliği</div>
                <div class="flow-text">Kusurlar karar destek amacıyla sıralanır; kesin uçuş kararı verilmez.</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_risk_legend() -> None:
    st.markdown(
        """
        <div class="risk-strip">
            <div class="risk-pill risk-low">Düşük: Risk skoru &lt; 0,50</div>
            <div class="risk-pill risk-medium">Orta: 0,50 – &lt; 3,00</div>
            <div class="risk-pill risk-high">Yüksek: 3,00 – &lt; 15,00</div>
            <div class="risk-pill risk-critical">Kritik: ≥ 15,00</div>
            <div class="risk-pill risk-note">Ömür çıktısı: göreceli öncelik göstergesi</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def clean_column_names(columns) -> list[str]:
    """Strip and make dataframe column names unique for Streamlit display."""

    seen: dict[str, int] = {}
    cleaned: list[str] = []
    for index, column in enumerate(columns, start=1):
        name = str(column).strip()
        if not name or name.lower().startswith("unnamed:"):
            name = f"column_{index}"
        count = seen.get(name, 0) + 1
        seen[name] = count
        cleaned.append(name if count == 1 else f"{name}_{count}")
    return cleaned


def get_spreadsheet_sheet_names(uploaded_file) -> list[str]:
    if uploaded_file is None:
        return []
    name = uploaded_file.name.lower()
    if name.endswith(".ods"):
        return read_ods_sheet_names(uploaded_file)
    if name.endswith((".xlsx", ".xls")):
        excel_file = pd.ExcelFile(uploaded_file)
        uploaded_file.seek(0)
        return excel_file.sheet_names
    return []


def main_risk_sheet_options(sheet_names: list[str]) -> list[str]:
    options = [
        sheet
        for sheet in sheet_names
        if not is_fracture_sheet_name(sheet) and not is_helper_sheet_name(sheet)
    ]
    return options or sheet_names


def is_fracture_sheet_name(sheet_name: str) -> bool:
    normalized = sheet_name.lower().replace(" ", "_")
    return "fracture" in normalized or "kirilma" in normalized or "kırılma" in normalized


def is_helper_sheet_name(sheet_name: str) -> bool:
    normalized = sheet_name.lower().replace(" ", "_")
    return normalized in {"assumptions", "grafikler", "charts", "graph", "graphs"}


def get_excel_defaults(uploaded_file) -> dict[str, float]:
    defaults = {
        "geometry_factor": 1.12,
        "delta_sigma": 250.0,
        "paris_c": 1.0e-10,
        "paris_m": 3.0,
        "critical_crack_size_mm": 5.0,
    }
    if uploaded_file is None or not uploaded_file.name.lower().endswith((".xlsx", ".xls")):
        return defaults

    try:
        assumptions = pd.read_excel(uploaded_file, sheet_name="Assumptions", header=None)
        uploaded_file.seek(0)
    except Exception:
        uploaded_file.seek(0)
        return defaults

    defaults["geometry_factor"] = _read_optional_float(assumptions, 12, 1, defaults["geometry_factor"])
    defaults["delta_sigma"] = _read_optional_float(assumptions, 13, 1, defaults["delta_sigma"])
    defaults["paris_c"] = _read_optional_float(assumptions, 16, 1, defaults["paris_c"])
    defaults["paris_m"] = _read_optional_float(assumptions, 17, 1, defaults["paris_m"])
    return defaults


def _read_optional_float(frame: pd.DataFrame, row: int, col: int, default: float) -> float:
    try:
        value = frame.iloc[row, col]
        if isinstance(value, str) and value.strip().startswith("="):
            if value.strip().upper() == "=10^(-10)":
                return 1.0e-10
            return default
        return float(value)
    except Exception:
        return default


def load_uploaded_file(uploaded_file, sheet_name: str | None = None) -> pd.DataFrame:
    """Load CSV or Excel input from Streamlit uploader."""

    if uploaded_file is None:
        return make_sample_dataframe()

    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        frame = pd.read_csv(uploaded_file)
    elif name.endswith(".ods"):
        frame = read_ods(uploaded_file, sheet_name=sheet_name)
    elif name.endswith((".xlsx", ".xls")):
        frame = pd.read_excel(uploaded_file, sheet_name=sheet_name)
    else:
        raise ValueError("Only CSV, Excel, and ODS files are supported.")

    frame.columns = clean_column_names(frame.columns)
    return frame


def guess_column(columns: list[str], candidates: list[str]) -> int:
    normalized = {
        column.lower().replace(" ", "").replace("-", "_"): index
        for index, column in enumerate(columns)
    }
    for candidate in candidates:
        key = candidate.lower().replace(" ", "").replace("-", "_")
        if key in normalized:
            return normalized[key]
    return 0


def optional_column_select(label: str, columns: list[str], candidates: list[str]) -> str:
    options = ["Yok / kullanılmayacak"] + columns
    guessed_index = guess_column(columns, candidates) + 1
    return st.selectbox(label, options, index=guessed_index)


def make_blank_input_rows() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Wrinkle_ID": f"W-{index:03d}",
                "Length_mm": None,
                "Width_mm": None,
                "Depth_mm": None,
                "Height_mm": None,
                "Delta_sigma_MPa": 250.0,
                "Geometry_factor_Y": 1.12,
                "Paris_Law_C": 1.0e-10,
                "Paris_Law_m": 3.0,
                "Critical_crack_size_mm": 5.0,
            }
            for index in range(1, 6)
        ]
    )


def priority_score_column(data: pd.DataFrame) -> str | None:
    for column in ("Severity_Index", "Risk_Skor", "risk_score"):
        if column in data.columns:
            return column
    return None


def sorted_priority_table(data: pd.DataFrame) -> pd.DataFrame:
    score_column = priority_score_column(data)
    if score_column is None:
        return data.reset_index(drop=True)
    return data.sort_values(score_column, ascending=False, na_position="last").reset_index(drop=True)


def prepare_display_dataframe(data: pd.DataFrame, include_all_columns: bool = False) -> pd.DataFrame:
    frame = data.copy()
    frame = frame.drop(columns=[column for column in DETAIL_ONLY_COLUMNS if column in frame.columns])
    if "risk_class" in frame.columns:
        frame["risk_class"] = frame["risk_class"].apply(normalize_risk_label)
    for final_column in ("Final Risk_Skor", "Final_Risk_Level", "Risk_Level"):
        if final_column in frame.columns:
            frame[final_column] = frame[final_column].apply(normalize_risk_label)
    if "effective_measure_source" in frame.columns:
        frame["effective_measure_source"] = frame["effective_measure_source"].replace(
            {
                "depth": "Depth kullanıldı",
                "height/2": "Height / 2 kullanıldı",
                "missing": "Eksik",
            }
        )

    if include_all_columns:
        ordered_columns = [column for column in RESULT_COLUMNS if column in frame.columns]
        remaining_columns = [column for column in frame.columns if column not in ordered_columns]
        frame = frame[ordered_columns + remaining_columns]

    frame = frame.rename(columns=DISPLAY_COLUMN_NAMES)
    frame.columns = clean_column_names(frame.columns)
    return frame.reset_index(drop=True)


def style_risk_table(data: pd.DataFrame):
    risk_columns = [column for column in ("Geometric_Risk", "Final_Risk_Level", "Risk_Level") if column in data.columns]
    if not risk_columns:
        return data

    def color_risk_cells(value: object) -> str:
        return RISK_CELL_COLORS.get(normalize_risk_label(value), "")

    return data.style.map(color_risk_cells, subset=risk_columns).format(
        {
            "da/dN (m/cycle)": "{:.3E}",
            "Kalan çevrim": "{:.0f}",
            "Estimated Life Cycles": "{:.0f}",
        },
        precision=4,
        na_rep="",
    )


def dataframe_to_csv(data: pd.DataFrame) -> bytes:
    buffer = StringIO()
    prepare_display_dataframe(data, include_all_columns=True).to_csv(buffer, index=False)
    return buffer.getvalue().encode("utf-8-sig")


def dataframe_to_excel(data: pd.DataFrame, fracture_layer: pd.DataFrame | None = None) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        export_data = prepare_display_dataframe(data, include_all_columns=True)
        export_data.to_excel(writer, index=False, sheet_name="Risk_Results")
        worksheet = writer.sheets["Risk_Results"]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F2937")

        risk_col_indexes = []
        for index, cell in enumerate(worksheet[1], start=1):
            if cell.value in {"Geometric_Risk", "Final_Risk_Level", "Risk_Level"}:
                risk_col_indexes.append(index)

        if risk_col_indexes:
            for row in range(2, worksheet.max_row + 1):
                for risk_col_index in risk_col_indexes:
                    cell = worksheet.cell(row=row, column=risk_col_index)
                    fill = RISK_EXCEL_FILLS.get(normalize_risk_label(cell.value))
                    if fill is not None:
                        cell.fill = fill

        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 45)

        for index, cell in enumerate(worksheet[1], start=1):
            if cell.value == "da/dN (m/cycle)":
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(row=row, column=index).number_format = "0.00E+00"
            if cell.value == "Kalan çevrim":
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(row=row, column=index).number_format = "0"

        detail_columns = [
            column
            for column in [
                "Wrinkle_ID",
                "Row_ID",
                "yorum_özeti",
                "geometrik_değerlendirme",
                "kırılma_mekaniği_değerlendirmesi",
                "ömür_öncelik_yorumu",
            ]
            if column in data.columns
        ]
        if detail_columns:
            detail_data = data[detail_columns].rename(columns=DISPLAY_COLUMN_NAMES)
            detail_data.to_excel(writer, index=False, sheet_name="Detail_Comments")
            detail_sheet = writer.sheets["Detail_Comments"]
            detail_sheet.freeze_panes = "A2"
            detail_sheet.auto_filter.ref = detail_sheet.dimensions
            for cell in detail_sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F2937")
            for column_cells in detail_sheet.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                detail_sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 16), 70)
        if fracture_layer is not None and not fracture_layer.empty:
            fracture_export = prepare_display_dataframe(fracture_layer, include_all_columns=True)
            fracture_export.to_excel(writer, index=False, sheet_name="Fracture_Analysis")
            fracture_sheet = writer.sheets["Fracture_Analysis"]
            fracture_sheet.freeze_panes = "A2"
            fracture_sheet.auto_filter.ref = fracture_sheet.dimensions
            for cell in fracture_sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F2937")
            for index, cell in enumerate(fracture_sheet[1], start=1):
                if cell.value in {"Geometric_Risk", "Final_Risk_Level", "Risk_Level"}:
                    for row in range(2, fracture_sheet.max_row + 1):
                        risk_cell = fracture_sheet.cell(row=row, column=index)
                        fill = RISK_EXCEL_FILLS.get(normalize_risk_label(risk_cell.value))
                        if fill is not None:
                            risk_cell.fill = fill
                if cell.value == "da/dN (m/cycle)":
                    for row in range(2, fracture_sheet.max_row + 1):
                        fracture_sheet.cell(row=row, column=index).number_format = "0.00E+00"
                if cell.value == "Kalan çevrim":
                    for row in range(2, fracture_sheet.max_row + 1):
                        fracture_sheet.cell(row=row, column=index).number_format = "0"
            for column_cells in fracture_sheet.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                fracture_sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 45)
    return buffer.getvalue()


def turkish_risk_counts(data: pd.DataFrame) -> pd.DataFrame:
    counts = data["risk_class"].map(RISK_DISPLAY_LABELS).value_counts(dropna=False)
    return counts.rename_axis("Risk sınıfı").to_frame("Satır sayısı")


def summary_text(summary: dict[str, object]) -> str:
    top_5 = summary.get("top_5_priority", [])
    top_items = []
    if isinstance(top_5, list):
        for item in top_5:
            identifier = item.get("Wrinkle_ID") or item.get("defect_id") or item.get("Row_ID")
            risk = item.get("Severity_Index") or item.get("Risk_Skor") or item.get("risk_score")
            top_items.append(f"{identifier} (risk={risk:.4g})" if isinstance(risk, (int, float)) else str(identifier))

    return "\n".join(
        [
            f"- Toplam kusur sayısı: {summary.get('row_count', 0)}",
            f"- Hesaplanabilen satır sayısı: {summary.get('complete_for_risk_count', 0)}",
            f"- Eksik/geçersiz satır sayısı: {summary.get('incomplete_or_invalid_count', 0)}",
            f"- En yüksek riskli kusur: {summary.get('max_risk_defect', '-')}",
            f"- En yüksek Delta K değerine sahip kusur: {summary.get('max_delta_k_defect', '-')}",
            f"- En düşük kalan çevrime sahip kusur: {summary.get('min_life_defect', '-')}",
            f"- Öncelikli incelenecek ilk 5 kusur: {', '.join(top_items) if top_items else '-'}",
        ]
    )


def find_id_column(data: pd.DataFrame) -> str | None:
    for column in ("Wrinkle_ID", "defect_id", "Row_ID"):
        if column in data.columns:
            return column
    return None


def build_fracture_layer(primary_results: pd.DataFrame, uploaded_fracture: pd.DataFrame) -> pd.DataFrame:
    if not uploaded_fracture.empty:
        frame = uploaded_fracture.copy()
        risk_column = "Risk_Level" if "Risk_Level" in frame.columns else "risk_class" if "risk_class" in frame.columns else None
        if risk_column:
            normalized_risk = frame[risk_column].apply(normalize_risk_label)
            frame = frame[normalized_risk.isin(["MEDIUM", "HIGH", "CRITICAL"])].copy()
            frame["risk_class"] = normalized_risk.loc[frame.index]
        if "Effective_Size" in frame.columns and "Effective_Size_mm" not in frame.columns:
            frame["Effective_Size_mm"] = frame["Effective_Size"]
        if "effective_measure" in frame.columns and "Effective_Size_mm" not in frame.columns:
            frame["Effective_Size_mm"] = frame["effective_measure"]
        if "Assumed_Crack_mm" not in frame.columns:
            if "Assumed_Crack" in frame.columns:
                frame["Assumed_Crack_mm"] = frame["Assumed_Crack"]
            elif "Effective_Size_mm" in frame.columns:
                frame["Assumed_Crack_mm"] = pd.to_numeric(frame["Effective_Size_mm"], errors="coerce") * 0.5
        if "Assumed_Crack_m" not in frame.columns and "Assumed_Crack_mm" in frame.columns:
            frame["Assumed_Crack_m"] = pd.to_numeric(frame["Assumed_Crack_mm"], errors="coerce") / 1000.0
        if "Geometry_Factor_Y" not in frame.columns:
            frame["Geometry_Factor_Y"] = 1.12
        if "Assumed_Stress_MPa" not in frame.columns:
            frame["Assumed_Stress_MPa"] = 250.0
        if "Paris_C" not in frame.columns:
            frame["Paris_C"] = 1.0e-10
        if "Paris_m" not in frame.columns:
            frame["Paris_m"] = 3.0
        if "Assumed_Crack_m" in frame.columns:
            frame["Delta_K_MPa_sqrt_m"] = (
                pd.to_numeric(frame["Geometry_Factor_Y"], errors="coerce")
                * pd.to_numeric(frame["Assumed_Stress_MPa"], errors="coerce")
                * (np.pi * pd.to_numeric(frame["Assumed_Crack_m"], errors="coerce")) ** 0.5
            )
        elif "Delta_K" in frame.columns and "Delta_K_MPa_sqrt_m" not in frame.columns:
            frame["Delta_K_MPa_sqrt_m"] = frame["Delta_K"]
        if "da_dN" in frame.columns and "da_dN_m_per_cycle" not in frame.columns:
            frame["da_dN_m_per_cycle"] = frame["da_dN"]
        elif "Paris_C" in frame.columns and "Paris_m" in frame.columns and "Delta_K_MPa_sqrt_m" in frame.columns:
            frame["da_dN_m_per_cycle"] = (
                pd.to_numeric(frame["Paris_C"], errors="coerce")
                * pd.to_numeric(frame["Delta_K_MPa_sqrt_m"], errors="coerce")
                ** pd.to_numeric(frame["Paris_m"], errors="coerce")
            )
        if "Estimated_Life_Cycles" not in frame.columns:
            if "Estimated_Life" in frame.columns:
                frame["Estimated_Life_Cycles"] = frame["Estimated_Life"]
            elif "da_dN_m_per_cycle" in frame.columns:
                dadn = pd.to_numeric(frame["da_dN_m_per_cycle"], errors="coerce")
                frame["Estimated_Life_Cycles"] = np.where(dadn > 0, 1.0 / dadn, np.nan)
        if "Estimated_Life" in frame.columns and "remaining_life_cycles" not in frame.columns:
            frame["remaining_life_cycles"] = frame["Estimated_Life"]
        elif "Estimated_Life_Cycles" in frame.columns and "remaining_life_cycles" not in frame.columns:
            frame["remaining_life_cycles"] = frame["Estimated_Life_Cycles"]
        if "Engineering_Comment" in frame.columns and "yorum_özeti" not in frame.columns:
            frame["yorum_özeti"] = frame["Engineering_Comment"]
        frame = add_inspection_priority(frame)
        return frame.reset_index(drop=True)

    if "risk_class" not in primary_results.columns:
        return pd.DataFrame()
    frame = primary_results[
        primary_results["risk_class"].apply(normalize_risk_label).isin(["MEDIUM", "HIGH", "CRITICAL"])
    ].reset_index(drop=True)
    return add_inspection_priority(frame)


def add_inspection_priority(data: pd.DataFrame) -> pd.DataFrame:
    frame = data.copy()
    life_column = None
    for column in ("Estimated_Life_Cycles", "Estimated_Life", "remaining_life_cycles"):
        if column in frame.columns:
            life_column = column
            break
    if life_column is None:
        frame["inspection_priority"] = "Not evaluated"
        return frame

    life = pd.to_numeric(frame[life_column], errors="coerce")
    frame["inspection_priority"] = pd.cut(
        life,
        bins=[float("-inf"), 5.0e6, 2.0e7, float("inf")],
        labels=["Immediate Review", "Scheduled Inspection", "Routine Monitoring"],
        right=False,
    ).astype("object")
    frame.loc[life.isna(), "inspection_priority"] = "Not evaluated"
    return frame


def fracture_summary_metrics(fracture_layer: pd.DataFrame) -> dict[str, object]:
    if fracture_layer.empty:
        return {
            "row_count": 0,
            "immediate_review": 0,
            "max_delta_k": None,
            "min_life": None,
        }

    priority = fracture_layer.get("inspection_priority", pd.Series(dtype="object")).astype(str)
    delta_k = pd.to_numeric(fracture_layer.get("Delta_K_MPa_sqrt_m", pd.Series(dtype="float64")), errors="coerce")
    life = pd.to_numeric(
        fracture_layer.get("Estimated_Life_Cycles", fracture_layer.get("remaining_life_cycles", pd.Series(dtype="float64"))),
        errors="coerce",
    )
    return {
        "row_count": int(len(fracture_layer)),
        "immediate_review": int((priority == "Immediate Review").sum()),
        "max_delta_k": float(delta_k.max()) if delta_k.notna().any() else None,
        "min_life": float(life.min()) if life.notna().any() else None,
    }


def normalize_risk_label(value: object) -> str:
    return normalize_risk_label_core(value)


def classify_severity(value: object) -> str:
    if pd.isna(value):
        return "INCOMPLETE"
    try:
        score = float(value)
    except Exception:
        return normalize_risk_label(value)
    if score < 0.50:
        return "LOW"
    if score < 3.00:
        return "MEDIUM"
    if score < 15.00:
        return "HIGH"
    return "CRITICAL"


def row_label(row: pd.Series) -> str:
    parts = []
    for column in ("Wrinkle_ID", "Row_ID", "risk_class", "Risk_Level", "Final Risk_Skor"):
        if column in row and pd.notna(row[column]):
            value = RISK_DISPLAY_LABELS.get(row[column], row[column])
            parts.append(f"{column}: {value}")
    return " | ".join(parts) if parts else f"Satır {row.name}"


def render_detail_panel(row: pd.Series) -> None:
    st.subheader("Seçilen Kusur Detay Yorumu")
    st.markdown(f"**Yorum özeti:** {row.get('yorum_özeti', '-')}")

    col1, col2, col3 = st.columns(3)
    col1.metric("Açısal sapma", _format_number(row.get("theta_degree"), "{:.4g}"))
    col2.metric("Delta K", _format_number(row.get("Delta_K_MPa_sqrt_m"), "{:.4g}"))
    col3.metric("Göreceli ömür", _format_number(row.get("relative_life_index"), "{:.1f}"))

    with st.container(border=True):
        st.markdown("#### 1. Geometrik Değerlendirme")
        st.write(row.get("geometrik_değerlendirme", "Geometrik değerlendirme oluşturulamadı."))

    with st.container(border=True):
        st.markdown("#### 2. Kırılma Mekaniği Değerlendirmesi")
        st.write(row.get("kırılma_mekaniği_değerlendirmesi", "Kırılma mekaniği değerlendirmesi oluşturulamadı."))

    with st.container(border=True):
        st.markdown("#### 3. Ömür / Öncelik Yorumu")
        st.write(row.get("ömür_öncelik_yorumu", "Ömür ve öncelik yorumu oluşturulamadı."))


def render_fracture_detail_panel(row: pd.Series) -> None:
    if "geometrik_değerlendirme" in row:
        render_detail_panel(row)
        return

    st.subheader("Seçilen Kusur Kırılma Analizi")
    st.markdown(f"**Risk seviyesi:** {row.get('Risk_Level', row.get('risk_class', '-'))}")

    col1, col2, col3 = st.columns(3)
    col1.metric("Açısal sapma", _format_number(row.get("Angular_Deviation"), "{:.4g}"))
    delta_k_value = row.get("Delta_K_MPa_sqrt_m", row.get("Delta_K"))
    dadn_value = row.get("da_dN_m_per_cycle", row.get("da_dN"))
    col2.metric("Delta K", _format_number(delta_k_value, "{:.4g}"))
    life_value = row.get("Estimated_Life_Cycles", row.get("Estimated_Life"))
    col3.metric("Paris ters-hız göstergesi", _format_number(life_value, "{:.0f}"))
    st.markdown(f"**Inspection Priority:** {row.get('inspection_priority', '-')}")

    with st.container(border=True):
        st.markdown("#### 1. Geometrik Değerlendirme")
        st.write(
            "Bu kusur ana risk katmanında Orta, Yüksek veya Kritik seviyeye girdiği için kırılma mekaniği katmanına aktarılmıştır. "
            f"Açısal sapma değeri {_format_number(row.get('Angular_Deviation'), '{:.4g}')} ve severity indeksi "
            f"{_format_number(row.get('Severity_Index'), '{:.4g}')} olarak okunmuştur."
        )

    with st.container(border=True):
        st.markdown("#### 2. Kırılma Mekaniği Değerlendirmesi")
        st.write(
            f"Delta K {_format_number(delta_k_value, '{:.4g}')} ve da/dN "
            f"{_format_number(dadn_value, '{:.3E}')} seviyesindedir. Delta K arttıkça Paris Law bağıntısı nedeniyle "
            "çatlak ilerleme hızı üstel olarak artar."
        )

    with st.container(border=True):
        st.markdown("#### 3. Ömür / Öncelik Yorumu")
        comment = row.get("Engineering_Comment", "")
        st.write(
            f"Paris ters-hız göstergesi {_format_number(life_value, '{:.0f}')} olarak hesaplanmıştır. "
            f"Excel değerlendirmesi: {comment if pd.notna(comment) else 'yorum yok'}."
        )


def _format_number(value: object, fmt: str) -> str:
    if pd.isna(value):
        return "-"
    try:
        return fmt.format(float(value))
    except Exception:
        return str(value)


inject_app_styles()
render_hero()
render_model_flow()
render_risk_legend()

data_mode = st.radio("Veri giriş yöntemi", ["Excel/CSV yükle", "Elle veri gir"], horizontal=True)

try:
    fracture_input_data = pd.DataFrame()
    defaults = {
        "geometry_factor": 1.12,
        "delta_sigma": 250.0,
        "paris_c": 1.0e-10,
        "paris_m": 3.0,
        "critical_crack_size_mm": 5.0,
    }
    if data_mode == "Excel/CSV yükle":
        uploaded_file = st.file_uploader("Kusur verisi yükle", type=["csv", "xlsx", "xls", "ods"])
        defaults = get_excel_defaults(uploaded_file)
        sheet_name = None
        sheet_names = get_spreadsheet_sheet_names(uploaded_file)
        if sheet_names:
            main_sheet_names = main_risk_sheet_options(sheet_names)
            default_primary = (
                main_sheet_names.index("Primary_Risk_Assessment")
                if "Primary_Risk_Assessment" in main_sheet_names
                else 0
            )
            sheet_name = st.selectbox(
                "Ana risk sayfası",
                main_sheet_names,
                index=default_primary,
            )
            fracture_sheet = None
            if "Advanced_Fracture_Analysis" in sheet_names:
                fracture_sheet = "Advanced_Fracture_Analysis"
            elif "Fracture_Analysis" in sheet_names:
                fracture_sheet = "Fracture_Analysis"
            if fracture_sheet:
                st.success(f"Kırılma analizi sayfası otomatik yüklendi: {fracture_sheet}")
                st.caption("Bu yüzden yukarıdaki listede sadece ana risk sayfası görünür; fracture sayfası Kırılma Analizi sekmesindedir.")
                fracture_input_data = load_uploaded_file(uploaded_file, sheet_name=fracture_sheet)
        input_data = load_uploaded_file(uploaded_file, sheet_name=sheet_name)
    else:
        st.write("Wrinkle geometrisini ve ömür modeli parametrelerini aşağıdaki tabloya girin.")
        input_data = st.data_editor(
            make_blank_input_rows(),
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "Wrinkle_ID": st.column_config.TextColumn("Wrinkle ID"),
                "Length_mm": st.column_config.NumberColumn("Uzunluk mm", min_value=0.0),
                "Width_mm": st.column_config.NumberColumn("Genişlik mm", min_value=0.0),
                "Depth_mm": st.column_config.NumberColumn("Derinlik mm", min_value=0.0),
                "Height_mm": st.column_config.NumberColumn("Yükseklik mm", min_value=0.0),
                "Delta_sigma_MPa": st.column_config.NumberColumn("Delta sigma MPa", min_value=0.0),
                "Geometry_factor_Y": st.column_config.NumberColumn("Y faktörü", min_value=0.0),
                "Paris_Law_C": st.column_config.NumberColumn("Paris C", min_value=0.0, format="%.12e"),
                "Paris_Law_m": st.column_config.NumberColumn("Paris m", min_value=0.0),
                "Critical_crack_size_mm": st.column_config.NumberColumn("Kritik çatlak boyutu mm", min_value=0.0),
            },
        )
except Exception as exc:  # pragma: no cover - Streamlit display path
    st.error(str(exc))
    st.stop()

if input_data.empty or len(input_data.columns) == 0:
    st.warning("Seçilen dosya veya Excel sayfası okunabilir tablo verisi içermiyor.")
    st.stop()

with st.sidebar:
    st.header("Kolon Seçimi")
    available_columns = [str(column) for column in input_data.columns]
    length_col = st.selectbox(
        "Uzunluk kolonu",
        available_columns,
        index=guess_column(available_columns, ["length", "length_mm", "Length_mm"]),
    )
    width_selection = st.selectbox(
        "Genişlik kolonu",
        available_columns,
        index=guess_column(available_columns, ["width", "width_mm", "Width_mm"]),
    )
    depth_selection = optional_column_select("Derinlik kolonu", available_columns, ["depth", "depth_mm", "Depth_mm"])
    height_selection = optional_column_select("Yükseklik kolonu", available_columns, ["height", "height_mm", "Height_mm"])
    width_col = width_selection
    depth_col = "" if depth_selection == "Yok / kullanılmayacak" else depth_selection
    height_col = "" if height_selection == "Yok / kullanılmayacak" else height_selection

    st.header("Risk Sınıfları")
    threshold_mode = st.radio("Eşik yöntemi", ["Açık severity eşikleri", "Manuel"], index=0)
    manual_thresholds = RiskThresholds()
    st.caption(
        "Varsayılan eşikler: <0,50 Düşük; 0,50–<3,00 Orta; "
        "3,00–<15,00 Yüksek; ≥15,00 Kritik."
    )
    if threshold_mode == "Manuel":
        medium_min = st.number_input("Orta risk başlangıcı", value=0.50, min_value=0.0)
        high_min = st.number_input("Yüksek risk başlangıcı", value=3.00, min_value=0.0)
        critical_min = st.number_input("Kritik risk başlangıcı", value=15.00, min_value=0.0)
        try:
            manual_thresholds = RiskThresholds(
                medium_min=medium_min,
                high_min=high_min,
                critical_min=critical_min,
            )
        except ValueError as exc:
            st.error(str(exc))
            st.stop()

    st.header("Ömür Modeli Parametreleri")
    compute_fracture_life = st.checkbox("Delta K, Paris Law ve kalan çevrim hesapla", value=True)
    delta_sigma_default = st.number_input(
        "Varsayılan Delta sigma (MPa)",
        value=float(defaults["delta_sigma"]),
        min_value=0.000001,
        format="%.6f",
    )
    geometry_factor_default = st.number_input(
        "Varsayılan Y faktörü",
        value=float(defaults["geometry_factor"]),
        min_value=0.000001,
        format="%.6f",
    )
    paris_c_default = st.number_input(
        "Varsayılan Paris C",
        value=float(defaults["paris_c"]),
        min_value=1.0e-20,
        format="%.12e",
    )
    paris_m_default = st.number_input(
        "Varsayılan Paris m",
        value=float(defaults["paris_m"]),
        min_value=0.000001,
        format="%.6f",
    )
    critical_crack_size_mm_default = st.number_input(
        "Varsayılan kritik çatlak boyutu (mm)",
        value=float(defaults["critical_crack_size_mm"]),
        min_value=0.000001,
        format="%.6f",
    )

columns = ColumnMap(length=length_col, width=width_col, depth=depth_col, height=height_col)
use_uploaded_fracture_layer = data_mode == "Excel/CSV yükle" and not fracture_input_data.empty
options = AnalysisOptions(
    columns=columns,
    risk_thresholds=manual_thresholds,
    compute_fracture_life=compute_fracture_life and not use_uploaded_fracture_layer,
    delta_sigma_default=delta_sigma_default,
    geometry_factor_default=geometry_factor_default,
    paris_c_default=paris_c_default,
    paris_m_default=paris_m_default,
    critical_crack_size_mm_default=critical_crack_size_mm_default,
)

try:
    result = run_analysis(input_data, options)
except MissingColumnError as exc:  # pragma: no cover - Streamlit display path
    st.error(str(exc))
    st.write("Yüklenen dosyada bulunan kolonlar:")
    st.dataframe(pd.DataFrame({"Kolon adı": input_data.columns}), use_container_width=True)
    st.stop()
except Exception as exc:  # pragma: no cover - Streamlit display path
    st.error(str(exc))
    st.stop()

data = result.data
if "Severity_Index" in data.columns:
    data["risk_class"] = data["Severity_Index"].apply(classify_severity)
elif "Risk_Skor" in data.columns:
    data["risk_class"] = data["Risk_Skor"].apply(classify_severity)
elif "Final Risk_Skor" in data.columns:
    data["risk_class"] = data["Final Risk_Skor"].apply(normalize_risk_label)
fracture_data = build_fracture_layer(data, fracture_input_data)
fracture_metrics = fracture_summary_metrics(fracture_data)
max_delta_k_value = result.summary.get("max_delta_k", 0.0)
if not fracture_data.empty and "Delta_K_MPa_sqrt_m" in fracture_data.columns:
    fracture_delta_k = pd.to_numeric(fracture_data["Delta_K_MPa_sqrt_m"], errors="coerce").dropna()
    if not fracture_delta_k.empty:
        max_delta_k_value = float(fracture_delta_k.max())

metric_cols = st.columns(5)
metric_cols[0].metric("Toplam satır", result.summary["row_count"])
metric_cols[1].metric("Hesaplanan satır", result.summary["complete_for_risk_count"])
metric_cols[2].metric("Eksik/geçersiz", result.summary["incomplete_or_invalid_count"])
metric_cols[3].metric("En yüksek risk", f"{result.summary.get('max_risk_score', 0.0):.4g}")
metric_cols[4].metric("En yüksek Delta K", f"{max_delta_k_value:.4g}")

st.markdown(
    """
    <div class="section-note">
        Ana sıralama Risk_Skor üzerinden yapılır. Genişlik ağırlıklı risk ana sınıflandırmada kullanılmamıştır;
        yalnızca duyarlılık göstergesi olarak tutulur. ΔK, Paris Yasası ve göreceli ömür çıktıları
        mühendislik kabullerine dayanır.
    </div>
    """,
    unsafe_allow_html=True,
)

results_tab, fracture_tab, summary_tab, charts_tab, quality_tab, input_tab = st.tabs(
    ["1. Ana Risk", "2. Kırılma Analizi", "3. Genel Özet", "4. Grafikler", "5. Veri Kontrolü", "6. Girdi Verisi"]
)

with results_tab:
    preferred_columns = [column for column in PRIMARY_RESULT_COLUMNS if column in data.columns]
    priority_table_data = sorted_priority_table(data[preferred_columns])
    st.subheader("Ana Risk Değerlendirmesi")
    st.markdown(
        """
        <div class="section-note">
            Bu katman wrinkle geometrisi, açısal sapma, severity/risk skorları ve risk sınıfını gösterir.
            Düşük kusurlar kırılma analizine aktarılmaz. Yükseklik/2 etkin ölçü yaklaşımı yalnızca derinlik yokken kullanılan mühendislik varsayımıdır.
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.dataframe(style_risk_table(prepare_display_dataframe(priority_table_data)), use_container_width=True)

    download_left, download_right = st.columns([1, 1])
    with download_left:
        st.download_button(
            "Sonuçları Excel olarak indir",
            dataframe_to_excel(data, fracture_data),
            file_name="wrinkle_risk_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with download_right:
        st.download_button(
            "Sonuçları CSV olarak indir",
            dataframe_to_csv(data),
            file_name="wrinkle_risk_results.csv",
            mime="text/csv",
            use_container_width=True,
        )

with fracture_tab:
    st.subheader("Kırılma Mekaniği Katmanı")
    fcols = st.columns(4)
    fcols[0].metric("Fracture satırı", fracture_metrics["row_count"])
    fcols[1].metric("Immediate Review", fracture_metrics["immediate_review"])
    fcols[2].metric("Maks. Delta K", _format_number(fracture_metrics["max_delta_k"], "{:.4g}"))
    fcols[3].metric("Min. göreceli gösterge", _format_number(fracture_metrics["min_life"], "{:.0f}"))
    st.markdown(
        """
        <div class="section-note">
            Bu katmanda yalnızca Orta, Yüksek ve Kritik kusurlar değerlendirilir. ΔK hesabında
            Y=1,12 ve σ=250 MPa varsayılan mühendislik kabulleri kullanılır; doğrulanmış sonlu elemanlar
            gerilme alanlarını temsil etmez. Göreceli ömür göstergesi mutlak kalan çevrim değildir.
        </div>
        """,
        unsafe_allow_html=True,
    )
    if fracture_data.empty:
        st.info("Kırılma analizi için uygun Orta, Yüksek veya Kritik kusur bulunamadı.")
    else:
        fracture_table_data = sorted_priority_table(fracture_data)
        st.dataframe(style_risk_table(prepare_display_dataframe(fracture_table_data, include_all_columns=True)), use_container_width=True)
        labels = [row_label(row) for _, row in fracture_table_data.iterrows()]
        selected_label = st.selectbox(
            "Kırılma detayı gösterilecek kusur",
            labels,
            help="Seçilen kusurun engineering comment ve detay yorumları aşağıdaki panelde gösterilir.",
        )
        selected_index = labels.index(selected_label)
        render_fracture_detail_panel(fracture_table_data.iloc[selected_index])

with summary_tab:
    st.subheader("Genel Özet Yorumu")
    st.markdown(
        """
        <div class="section-note">
            Özet bölümü, proje sunumunda hızlı anlatım için hazırlanmıştır: toplam veri kalitesi,
            en kritik kusurlar ve fracture öncelikleri tek ekranda okunur.
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(summary_text(result.summary))
    with st.expander("Validation & Future Improvements", expanded=True):
        st.markdown(
            """
            - Bu model screening-oriented karar destek aracıdır; certification veya uçuşa elverişlilik kararı vermez.
            - Ömür çıktıları mutlak servis ömrü değil, kusurlar arası göreceli önceliklendirme göstergesidir.
            - ΔK hesapları kabul edilen gerilme değerleriyle çalışır; doğrulanmış sonlu elemanlar gerilme alanı yerine geçmez.
            - Paris C ve Paris m değerleri laminat ve malzeme özelinde deneysel olarak kalibre edilmelidir.
            - Gelecek aşamada sonlu elemanlar entegrasyonu ve yorulma deneyi karşılaştırması eklenmelidir.
            """
        )
    top_columns = [
        column
        for column in [
            "Wrinkle_ID",
            "Row_ID",
            "risk_score",
            "risk_class",
            "Delta_K_MPa_sqrt_m",
            "remaining_life_cycles",
            "fracture_status",
            "yorum_özeti",
        ]
        if column in data.columns
    ]
    st.subheader("Öncelikli ilk 5 kusur")
    top_5 = sorted_priority_table(data[top_columns]).head(5)
    st.dataframe(style_risk_table(prepare_display_dataframe(top_5)), use_container_width=True)

with charts_tab:
    st.subheader("Görsel Analiz Paneli")
    st.caption("Grafikler risk dağılımını, Delta K seviyelerini ve öncelik sıralamasını hızlı okumak için kullanılır.")
    left, right = st.columns(2)
    with left:
        st.pyplot(plot_risk_distribution(data))
        st.pyplot(plot_delta_k_distribution(fracture_data if not fracture_data.empty else data))
        st.pyplot(plot_risk_class_counts(data))
    with right:
        st.pyplot(plot_theta_vs_measure(data))
        st.pyplot(plot_remaining_life_distribution(fracture_data if not fracture_data.empty else data))
        st.pyplot(plot_top_critical_defects(data, id_column=find_id_column(data)))

with quality_tab:
    st.subheader("Geometric Risk Dağılımı")
    st.dataframe(turkish_risk_counts(data), use_container_width=True)
    st.subheader("Veri Kalitesi Bayrakları")
    st.dataframe(
        data["data_quality_status"]
        .value_counts(dropna=False)
        .rename_axis("Veri durumu")
        .to_frame("Satır sayısı"),
        use_container_width=True,
    )
    st.dataframe(
        style_risk_table(prepare_display_dataframe(sorted_priority_table(data), include_all_columns=True)),
        use_container_width=True,
    )

with input_tab:
    st.dataframe(input_data, use_container_width=True)
